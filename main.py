import smtplib
import speech_recognition as sr
import pyttsx3
from email.message import EmailMessage
import openpyxl as xl

wb = xl.load_workbook('contacts.xlsx')
sheet = wb['Sheet1']
contact_list = {}
x = 2
for x in range(2, sheet.max_row + 1):
    cell1 = sheet.cell(x, 2)
    cell2 = sheet.cell(x, 3)
    contact_list[cell1.value] = cell2.value

listener = sr.Recognizer()

engine = pyttsx3.init()
engine.setProperty('rate', 180)  # reduces WPM to 180
engine.setProperty('voice', 'com.apple.speech.synthesis.voice.samantha.premium')  # This voice is best suited in macos


def talk(text):
    engine.say(text)
    engine.runAndWait()


def mike_out():
    try:
        with sr.Microphone() as source:
            print('listening...')
            voice = listener.listen(source)
            info = listener.recognize_google(voice)
            print(info)
            return info.lower()
    except:
        pass


def send_email(receiver, subject, message):
    server = smtplib.SMTP('smtp.gmail.com', 587)  # to connect to server
    server.starttls()  # tells that it is an secure connection
    # Make sure to give app access in your Google account
    server.login('gojo.testing123@gmail.com', 'hellogojo')
    email = EmailMessage()
    email['From'] = 'gojo.testing123@gmail.com'
    email['To'] = receiver
    email['Subject'] = subject
    email.set_content(message)
    server.send_message(email)


def create_email():
    print("Receiver's names(you can say 'All' for all your contacts): ", end=' ')
    talk('To Whom you want to send this Email.')
    talk('you can send this to multiple people by connecting with and!')
    l_names = mike_out()
    if l_names == 'quit':
        exit()
    names = l_names.split(' and ')
    for name in names:
        if name not in contact_list:
            names.remove(name)
    receivers = []
    if len(names) == 0:
        talk("sorry there are no similar email addresses in your contacts")
        talk("please try again later")
        exit()
    for name in names:
        receivers.append(contact_list[name])
    print("Receiver's Email addresses: ", *receivers)
    print('Subject of this Email: ', end=' ')
    talk('What is the subject of this Email?')
    subject = mike_out()
    if subject == 'quit':
        exit()
    print('Text in your Email: ', end=' ')
    talk('Tell me the text in your Email?')
    message = mike_out()
    if message == 'quit':
        exit()
    for receiver in receivers:
        send_email(receiver, subject, message)
    print('Email sent successfully to', *names)
    talk('Your emails has been sent successfully')

    talk('Do you want to send another email?')
    send_more = mike_out()
    if send_more == 'yes' or send_more == 'yep':
        new_or_not()


def new_or_not():
    print("Say 'YES' if it is known contact else say 'NO': ")
    talk('Are you sending this email to a know contact?')
    opt = mike_out()
    if opt == 'yes':
        create_email()
    elif opt == 'no':
        talk('type the name and address of new contact!')
        name = input("Type name of new contact to add: ")
        email = input("Type Email address of new contact to add: ")
        contact_list[name] = email
        new_cell0 = sheet.cell(x + 1, 1)
        new_cell1 = sheet.cell(x + 1, 2)
        new_cell2 = sheet.cell(x + 1, 3)
        new_cell0.value = x - 1
        new_cell1.value = name
        new_cell2.value = email
        wb.save('contacts.xlsx')
        create_email()


talk('hello! I am samantha your email bot')
talk("you can exit at any point by saying 'quit'")
new_or_not()
wb.close()
